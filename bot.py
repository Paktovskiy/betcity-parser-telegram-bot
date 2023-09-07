from telebot import TeleBot, apihelper
from time import sleep, mktime, time
from openpyxl import load_workbook
from datetime import datetime
from requests import get
from re import search
import fake_useragent
import random

TOKEN, LOGIN, PASSWORD = '', '', ''  # your telegram bot token, proxy login and password
main_channel, debug_channel = '', ''  # your telegram channel names
live = []  # your proxy IPs
banned, potentially_banned = [], []
scanned_games = set()
irrelevant_champs = 'Статистика|Киберфутбол|7x7|7х7|6x6|5x5|3x3'
relevant_minutes = (15, 30, 45, 60, 75)
divisor = 2.25
bot = TeleBot(TOKEN)


def send_tg_message(channel, message, retry=3):
    proxy_tg = random.choice(live)
    apihelper.proxy = {'https': f'http://{LOGIN}:{PASSWORD}@{proxy_tg}'}
    try:
        bot.send_message(channel, message)
    except Exception as excp:
        sleep(random.uniform(10, 15))
        if retry:
            send_tg_message(channel, message, retry=(retry - 1))
        else:
            with open('log.txt', 'a', encoding="UTF-8") as f:
                date = datetime.fromtimestamp(int(time()))
                f.write(f'{date}\n{str(type(excp))[8:-2]}\n{message}\n\n')


send_tg_message(debug_channel, 'Bot has been started')
last_signal_time = last_xlsx_update_time = int(time())

while True:
    no_games = True

    url_1 = 'https://ad.betcity.ru/d/on_air/events'
    params_1 = (('rev', '5'), ('ids_sp', '1'), ('ver', '341'), ('csn', 'ooca9s'))
    headers_1 = {'user-agent': fake_useragent.UserAgent().random,
                 'authority': 'ad.betcity.ru',
                 'accept': 'application/json, text/plain, */*',
                 'accept-language': 'en-GB,en;q=0.9',
                 'content-type': 'application/x-www-form-urlencoded',
                 'origin': 'https://betcity.ru',
                 'referer': 'https://betcity.ru/',
                 'sec-fetch-dest': 'empty',
                 'sec-fetch-mode': 'cors',
                 'sec-fetch-site': 'same-site',
                 'connection': 'keep-alive'
                 }
    proxy_1 = random.choice(live)
    proxies_1 = {'https': f'http://{LOGIN}:{PASSWORD}@{proxy_1}'}

    try:
        response_1 = get(url_1, params=params_1, headers=headers_1, proxies=proxies_1, timeout=2)
        json_1 = response_1.json()
    except Exception as ex:
        potentially_banned.append(proxy_1)
        if potentially_banned.count(proxy_1) == 5:
            banned.append(live.pop(live.index(proxy_1)))
        send_tg_message(debug_channel, f'url_1\n{str(type(ex))[8:-2]}\n{proxy_1}')
        sleep(random.uniform(3, 6))
        continue

    for champ_id in json_1['reply']['sports']['1']['chmps']:
        champ_name = json_1['reply']['sports']['1']['chmps'][champ_id]['name_ch']

        if not search(irrelevant_champs, champ_name):
            no_games = False
            now = int(time())

            for game_id in json_1['reply']['sports']['1']['chmps'][champ_id]['evts']:
                time_name = json_1['reply']['sports']['1']['chmps'][champ_id]['evts'][game_id]['time_name']

                if time_name in ('1-й тайм', '2-й тайм'):
                    try:
                        first_half = json_1['reply']['sports']['1']['chmps'][champ_id]['evts'][game_id]['min']
                        kick_off = json_1['reply']['sports']['1']['chmps'][champ_id]['evts'][game_id]['md_min']
                        kick_off = mktime(datetime.strptime(kick_off, "%Y-%m-%d %H:%M:%S").timetuple())
                        current_min = int(first_half + (now - kick_off) // 60) + 1
                    except KeyError:
                        continue
                else:
                    continue

                game_log = f'{game_id} {current_min}'

                if current_min in relevant_minutes and game_log not in scanned_games:
                    sleep(random.uniform(3, 6))

                    url_2 = 'https://ad.betcity.ru/d/on_air/bets'
                    params_2 = (('rev', '8'), ('add', 'dep_event'), ('ids', game_id),
                                ('stat', '1'), ('ver', '341'), ('csn', 'ooca9s'))
                    headers_2 = {'user-agent': fake_useragent.UserAgent().random,
                                 'authority': 'ad.betcity.ru',
                                 'accept': 'application/json, text/plain, */*',
                                 'accept-language': 'en-GB,en;q=0.9',
                                 'content-type': 'application/x-www-form-urlencoded',
                                 'origin': 'https://betcity.ru',
                                 'referer': 'https://betcity.ru/',
                                 'sec-fetch-dest': 'empty',
                                 'sec-fetch-mode': 'cors',
                                 'sec-fetch-site': 'same-site',
                                 'connection': 'keep-alive'
                                 }
                    proxy_2 = random.choice(live)
                    proxies_2 = {'https': f'http://{LOGIN}:{PASSWORD}@{proxy_2}'}

                    try:
                        response_2 = get(url_2, params=params_2, headers=headers_2, proxies=proxies_2, timeout=2)
                        json_2 = response_2.json()
                    except Exception as ex:
                        potentially_banned.append(proxy_2)
                        if potentially_banned.count(proxy_2) == 5:
                            banned.append(live.pop(live.index(proxy_2)))
                        send_tg_message(debug_channel, f'url_2\n{str(type(ex))[8:-2]}\n{proxy_2}')
                        continue

                    try:
                        ht_on_goal = json_2['reply']['live_stat'][game_id]['tbl']['1']['sg']
                        at_on_goal = json_2['reply']['live_stat'][game_id]['tbl']['2']['sg']
                        total_on_goal = ht_on_goal + at_on_goal
                        message_on_goal = f'{total_on_goal} ({ht_on_goal}:{at_on_goal})'
                    except (KeyError, TypeError):
                        scanned_games.add(game_log)
                        continue

                    rule_1 = rule_2 = rule_3 = False

                    if total_on_goal >= int(current_min / divisor):
                        rule_1, rule_tag = True, '#1'
                    elif (ht_on_goal >= 10 and at_on_goal <= 1) or (ht_on_goal <= 1 and at_on_goal >= 10):
                        rule_2, rule_tag = True, '#2'
                    elif abs(ht_on_goal - at_on_goal) >= 12:
                        rule_3, rule_tag = True, '#3'

                    if rule_1 or rule_2 or rule_3:
                        home_team = json_2['reply']['sports']['1']['chmps'][champ_id]['evts'][game_id]['name_ht']
                        away_team = json_2['reply']['sports']['1']['chmps'][champ_id]['evts'][game_id]['name_at']
                        score = json_2['reply']['sports']['1']['chmps'][champ_id]['evts'][game_id]['sc_ev']

                        try:
                            ht_on_target = json_2['reply']['live_stat'][game_id]['tbl']['1']['sh']
                            at_on_target = json_2['reply']['live_stat'][game_id]['tbl']['2']['sh']
                            total_on_target = ht_on_target + at_on_target
                            message_on_target = f'{total_on_target} ({ht_on_target}:{at_on_target})'
                        except (KeyError, TypeError):
                            message_on_target = 'нет данных'

                        try:
                            ht_corners = json_2['reply']['live_stat'][game_id]['tbl']['1']['c']
                            at_corners = json_2['reply']['live_stat'][game_id]['tbl']['2']['c']
                            total_corners = ht_corners + at_corners
                            message_corners = f'{total_corners} ({ht_corners}:{at_corners})'
                        except (KeyError, TypeError):
                            message_corners = 'нет данных'

                        try:
                            ht_goal_kicks = json_2['reply']['live_stat'][game_id]['tbl']['1']['gk']
                            at_goal_kicks = json_2['reply']['live_stat'][game_id]['tbl']['2']['gk']
                            total_goal_kicks = ht_goal_kicks + at_goal_kicks
                            message_goal_kicks = f'{total_goal_kicks} ({ht_goal_kicks}:{at_goal_kicks})'
                        except (KeyError, TypeError):
                            message_goal_kicks = 'нет данных'

                        try:
                            total_over = json_2['reply']['sports']['1']['chmps'][champ_id]['evts'][game_id]['main'][
                                '72']['data'][game_id]['blocks']['T1m']['Tot']
                            to_odds = json_2['reply']['sports']['1']['chmps'][champ_id]['evts'][game_id]['main'][
                                '72']['data'][game_id]['blocks']['T1m']['Tb']['kf']
                            line = f'ТБ {total_over} –> {to_odds}'
                        except KeyError:
                            line, total_over = 'Приём пари временно остановлен', '–'

                        signal = f'{rule_tag}\n' \
                                 '\n' \
                                 f'{champ_name[8:]}\n' \
                                 f'{home_team} – {away_team}\n' \
                                 f'Счет: {score}\n' \
                                 f'Время: {current_min}\n' \
                                 f'Удары: {message_on_goal}\n' \
                                 f'В створ: {message_on_target}\n' \
                                 f'Угловые: {message_corners}\n' \
                                 f'Удары от ворот: {message_goal_kicks}\n' \
                                 '\n' \
                                 f'https://betcity.ru/ru/live/soccer/{champ_id}/{game_id}\n' \
                                 f'{line}'

                        send_tg_message(main_channel, signal)
                        last_signal_time = int(time())

                        book = load_workbook('db.xlsx')
                        sheet = book.active
                        sheet.append([rule_tag, f'{champ_name[8:]}', f'{home_team} – {away_team}', score, current_min,
                                      message_on_goal, message_on_target, message_corners, message_goal_kicks,
                                      line, '', '', champ_id, game_id, total_over])
                        book.save('db.xlsx')
                        book.close()

                    scanned_games.add(game_log)

    if no_games:
        scanned_games.clear()
        cooldown = random.uniform(595, 605)
        if int(time()) - last_xlsx_update_time >= 12 * 60 * 60:
            cooldown = random.uniform(13, 17)
            send_tg_message(debug_channel, 'Updating database...')
            book = load_workbook('db.xlsx')
            sheet = book.active
            for row in range(2, sheet.max_row + 1):
                if sheet[row][10].value is None:
                    champ_id_db, game_id_db = sheet[row][12].value, sheet[row][13].value
                    sleep(random.uniform(3, 6))

                    url_3 = f'https://ad.betcity.ru/d/result/event'
                    params_3 = (('rev', '1'), ('id', game_id_db), ('ver', '341'), ('csn', 'ooca9s'))
                    headers_3 = {'user-agent': fake_useragent.UserAgent().random,
                                 'authority': 'ad.betcity.ru',
                                 'accept': 'application/json, text/plain, */*',
                                 'accept-language': 'en-GB,en;q=0.9',
                                 'content-type': 'application/x-www-form-urlencoded',
                                 'origin': 'https://betcity.ru',
                                 'referer': 'https://betcity.ru/',
                                 'sec-fetch-dest': 'empty',
                                 'sec-fetch-mode': 'cors',
                                 'sec-fetch-site': 'same-site',
                                 'connection': 'keep-alive'
                                 }
                    proxy_3 = random.choice(live)
                    proxies_3 = {'https': f'http://{LOGIN}:{PASSWORD}@{proxy_3}'}

                    try:
                        response_3 = get(url_3, params=params_3, headers=headers_3, proxies=proxies_3, timeout=2)
                        json_3 = response_3.json()
                    except Exception as ex:
                        potentially_banned.append(proxy_3)
                        if potentially_banned.count(proxy_3) == 5:
                            banned.append(live.pop(live.index(proxy_3)))
                        send_tg_message(debug_channel, f'url_3\n{str(type(ex))[8:-2]}\n{proxy_3}')
                        continue

                    try:
                        score_db = json_3['reply']['sports']['1']['chmps'][champ_id_db]['evts'][game_id_db]['sc_ev']
                        sheet[row][10].value = score_db
                        if sheet[row][14].value != '–':
                            if sum(map(int, score_db.split(':'))) > sheet[row][14].value:
                                sheet[row][11].value = 'Выигрыш'
                            else:
                                sheet[row][11].value = 'Проигрыш'
                        else:
                            sheet[row][11].value = '–'
                    except (KeyError, TypeError, ValueError):
                        sheet[row][10].value = sheet[row][11].value = 'Ошибка при получении данных'
                        continue
            book.save('db.xlsx')
            book.close()
            send_tg_message(debug_channel, 'Database is up to date')
            last_xlsx_update_time = last_signal_time = int(time())
    else:
        cooldown = random.uniform(13, 17)

    if int(time()) - last_signal_time >= 1800:
        if no_games:
            send_tg_message(debug_channel, 'Bot is online, no games')
        else:
            send_tg_message(debug_channel, 'Bot is online')
        last_signal_time = int(time())

    sleep(cooldown)
